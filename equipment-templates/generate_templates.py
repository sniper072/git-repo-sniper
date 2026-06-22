#!/usr/bin/env python3
"""Generate personnel equipment templates per organizational unit."""

from __future__ import annotations

import argparse
import re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
import sys

if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from template_engine import (
    find_template,
    load_template,
    render_template,
    write_download_test_package,
    write_template_bundle,
    write_template_document,
)

# Column aliases for auto-detection (Russian + English)
COLUMN_ALIASES = {
    "model": ["марка", "модель", "model", "наименование", "тип", "устройство", "оборудован"],
    "serial": ["серийный", "serial", "s/n", "sn", "сер. номер", "серийный номер"],
    "inventory": ["инвентар", "inventory", "инв. номер", "инвентарный", "инвентарный номер"],
    "location": ["местоположение", "место размещения", "место", "location", "расположение", "кабинет", "этаж", "адрес"],
    "unit": ["подразделение", "отдел", "unit", "департамент", "служба", "филиал", "место размещения", "офис"],
    "office": ["офис", "office"],
    "employee": ["сотрудник", "фио", "employee", "ответственный", "пользователь"],
    "equipment_name": ["наименование", "оборудование", "название", "device", "принтер"],
    "prev_counter": ["предыдущий месяц", "пред. месяц", "счетчик пред"],
    "report_counter": ["отчетный месяц", "отч. месяц", "счетчик отч"],
    "print_count": ["кол-во отпечатков", "отпечатков", "количество отпечатков", "принтов"],
    "row_number": ["№ п/п", "№ п\\п", "номер", "n п/п"],
}

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TABLE_HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def detect_column(columns: list[str], field: str) -> str | None:
    aliases = COLUMN_ALIASES[field]
    for column in columns:
        normalized = normalize_header(column)
        if any(alias in normalized for alias in aliases):
            return column
    return None


def load_txt_context(txt_dir: Path) -> str:
    if not txt_dir.exists():
        return ""

    parts: list[str] = []
    for txt_file in sorted(txt_dir.glob("*.txt")):
        content = txt_file.read_text(encoding="utf-8", errors="replace").strip()
        if content:
            parts.append(content)
    return "\n\n".join(parts)


def read_equipment_workbook(xlsx_path: Path) -> dict[str, pd.DataFrame]:
    workbook = pd.ExcelFile(xlsx_path)
    sheets: dict[str, pd.DataFrame] = {}

    for sheet_name in workbook.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str)
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            continue
        df.columns = [str(col).strip() for col in df.columns]
        sheets[sheet_name.strip() or "Без названия"] = df

    return sheets


def map_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    columns = list(df.columns)
    mapping = {
        "model": detect_column(columns, "model"),
        "serial": detect_column(columns, "serial"),
        "inventory": detect_column(columns, "inventory"),
        "location": detect_column(columns, "location"),
        "unit": detect_column(columns, "unit"),
        "employee": detect_column(columns, "employee"),
        "equipment_name": detect_column(columns, "equipment_name"),
        "prev_counter": detect_column(columns, "prev_counter"),
        "report_counter": detect_column(columns, "report_counter"),
        "print_count": detect_column(columns, "print_count"),
        "row_number": detect_column(columns, "row_number"),
    }

    result = pd.DataFrame()
    for field, source in mapping.items():
        if source:
            result[field] = df[source].fillna("").astype(str).str.strip()
        else:
            result[field] = ""

    if result["model"].eq("").all() and not result["equipment_name"].eq("").all():
        result["model"] = result["equipment_name"]

    placement_col = next(
        (column for column in columns if "место размещения" in normalize_header(column)),
        None,
    )
    office_col = detect_column(columns, "office")

    if office_col:
        result["office"] = df[office_col].fillna("").astype(str).str.strip()
    else:
        result["office"] = ""

    if placement_col and result["unit"].str.strip().eq("").all():
        result["unit"] = df[placement_col].fillna("").astype(str).str.strip()

    if placement_col:
        placement = df[placement_col].fillna("").astype(str).str.strip()
        office = (
            df[office_col].fillna("").astype(str).str.strip()
            if office_col
            else pd.Series([""] * len(df), index=df.index)
        )
        result["location"] = placement
        both = office.ne("") & placement.ne("")
        result.loc[both, "location"] = placement[both] + " (" + office[both] + ")"
        office_only = office.ne("") & placement.eq("")
        result.loc[office_only, "location"] = office[office_only]

    return result


def iter_devices(df: pd.DataFrame) -> list[pd.Series]:
    devices: list[pd.Series] = []
    for _, row in df.iterrows():
        if not str(row.get("model", "")).strip() and not str(row.get("serial", "")).strip():
            continue
        devices.append(row)
    return devices


def device_slug(device: pd.Series) -> str:
    inventory = sanitize_filename(str(device.get("inventory", "") or "no-inv")).replace("+", "plus")
    serial = sanitize_filename(str(device.get("serial", "") or "no-serial"))
    inventory = re.sub(r"[^\w\-]+", "_", inventory)
    serial = re.sub(r"[^\w\-]+", "_", serial)
    return f"{inventory}_{serial}"


def safe_text(value: object) -> str:
    text = str(value or "").strip()
    return text if text else "-"


def validate_xlsx(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    workbook.close()


def save_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    validate_xlsx(path)


def save_csv(path: Path, frame: pd.DataFrame) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, encoding="utf-8-sig")
    return path


CHECKLIST_ITEMS = [
    ("Модель соответствует учетным данным", "model"),
    ("Серийный номер соответствует", "serial"),
    ("Инвентарный номер соответствует", "inventory"),
    ("Место размещения соответствует", "location"),
    ("Офис соответствует", "office"),
    ("Подразделение соответствует", "unit"),
    ("Показания счетчика за предыдущий месяц зафиксированы", "prev_counter"),
    ("Показания счетчика за отчетный месяц зафиксированы", "report_counter"),
    ("Количество отпечатков за период проверено", "print_count"),
    ("Устройство включено и исправно", None),
    ("Печать / сканирование выполняется", None),
    ("Ответственный сотрудник определен", "employee"),
    ("Внешний вид и комплектность в норме", None),
    ("Замечания отсутствуют", None),
]


def device_card_dataframe(device: pd.Series) -> pd.DataFrame:
    rows = [
        ("Документ", "ИНВЕНТАРНАЯ КАРТОЧКА ОРГТЕХНИКИ"),
        ("Дата составления", date.today().strftime("%d.%m.%Y")),
        ("№ п/п", safe_text(device.get("row_number", ""))),
        ("Инвентарный номер", safe_text(device.get("inventory", ""))),
        ("Марка / модель", safe_text(device.get("model", ""))),
        ("Серийный номер", safe_text(device.get("serial", ""))),
        ("Место размещения", safe_text(device.get("location", ""))),
        ("Офис", safe_text(device.get("office", ""))),
        ("Подразделение", safe_text(device.get("unit", ""))),
        ("Счетчик (предыдущий месяц)", safe_text(device.get("prev_counter", ""))),
        ("Счетчик (отчетный месяц)", safe_text(device.get("report_counter", ""))),
        ("Кол-во отпечатков", safe_text(device.get("print_count", ""))),
        ("Ответственный", safe_text(device.get("employee", ""))),
        ("Состояние", "Исправно / Неисправно"),
        ("Примечание", ""),
        ("Материально ответственное лицо", ""),
        ("Инвентаризатор", ""),
    ]
    return pd.DataFrame(rows, columns=["Поле", "Значение"])


def device_checklist_dataframe(device: pd.Series) -> pd.DataFrame:
    rows: list[dict[str, object]] = [
        {
            "№": "",
            "Пункт проверки": "Инвентарный номер",
            "Да / Нет": "",
            "Фактическое значение": safe_text(device.get("inventory", "")),
            "Примечание": "",
        },
        {
            "№": "",
            "Пункт проверки": "Модель",
            "Да / Нет": "",
            "Фактическое значение": safe_text(device.get("model", "")),
            "Примечание": "",
        },
        {
            "№": "",
            "Пункт проверки": "Серийный номер",
            "Да / Нет": "",
            "Фактическое значение": safe_text(device.get("serial", "")),
            "Примечание": "",
        },
        {
            "№": "",
            "Пункт проверки": "Местоположение",
            "Да / Нет": "",
            "Фактическое значение": safe_text(device.get("location", "")),
            "Примечание": "",
        },
        {
            "№": "",
            "Пункт проверки": "Дата проверки",
            "Да / Нет": "",
            "Фактическое значение": date.today().strftime("%d.%m.%Y"),
            "Примечание": "",
        },
    ]

    for index, (label, field_key) in enumerate(CHECKLIST_ITEMS, start=1):
        fact = safe_text(device.get(field_key, "")) if field_key else ""
        rows.append(
            {
                "№": index,
                "Пункт проверки": label,
                "Да / Нет": "",
                "Фактическое значение": "" if fact == "-" else fact,
                "Примечание": "",
            }
        )

    rows.append(
        {
            "№": "",
            "Пункт проверки": "Проверил",
            "Да / Нет": "",
            "Фактическое значение": "",
            "Примечание": "",
        }
    )
    rows.append(
        {
            "№": "",
            "Пункт проверки": "Подпись ответственного",
            "Да / Нет": "",
            "Фактическое значение": "",
            "Примечание": "",
        }
    )
    return pd.DataFrame(rows, columns=["№", "Пункт проверки", "Да / Нет", "Фактическое значение", "Примечание"])


def write_label_value_block(ws, start_row: int, fields: list[tuple[str, str]], label_width: int = 2, value_width: int = 4) -> int:
    row = start_row
    for label, value in fields:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_width)
        ws.merge_cells(start_row=row, start_column=label_width + 1, end_row=row, end_column=label_width + value_width)
        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=label_width + 1, value=value or "—")
        style_cell(label_cell, bold=True, fill=TABLE_HEADER_FILL)
        style_cell(value_cell, wrap=True)
        row += 1
    return row


def write_device_card(output_path: Path, device: pd.Series, template_dir: Path) -> list[Path]:
    template = load_template(find_template("card", template_dir))
    rendered = render_template(template, device)
    write_template_document(output_path, "Card", rendered)
    txt_path = output_path.with_suffix(".txt")
    txt_path.write_text(rendered, encoding="utf-8")
    return [output_path, txt_path]


def write_device_checklist(output_path: Path, device: pd.Series, template_dir: Path) -> list[Path]:
    template = load_template(find_template("checklist", template_dir))
    rendered = render_template(template, device)
    write_template_document(output_path, "Checklist", rendered)
    txt_path = output_path.with_suffix(".txt")
    txt_path.write_text(rendered, encoding="utf-8")
    return [output_path, txt_path]


def write_device_daily_report(output_path: Path, device: pd.Series, template_dir: Path) -> list[Path]:
    template = load_template(find_template("daily_report", template_dir))
    rendered = render_template(template, device)
    write_template_document(output_path, "DailyReport", rendered)
    txt_path = output_path.with_suffix(".txt")
    txt_path.write_text(rendered, encoding="utf-8")
    return [output_path, txt_path]


def write_device_sample_bundle(
    output_path: Path, device: pd.Series, template_dir: Path
) -> Path:
    documents = {
        "Card": render_template(load_template(find_template("card", template_dir)), device),
        "Checklist": render_template(load_template(find_template("checklist", template_dir)), device),
        "DailyReport": render_template(load_template(find_template("daily_report", template_dir)), device),
    }
    write_template_bundle(output_path, documents)
    return output_path


def write_open_me_first(output_path: Path) -> Path:
    frame = pd.DataFrame(
        [
            {"Test": "OK", "Message": "If you can read this row, Excel download works"},
            {"Test": "1", "Message": "Open the Card and Checklist files next"},
        ]
    )
    save_workbook(output_path, {"Test": frame})
    save_csv(output_path.with_suffix(".csv"), frame)
    return output_path


def generate_device_documents(
    xlsx_path: Path,
    output_dir: Path,
    template_dir: Path,
    *,
    cards: bool = True,
    checklists: bool = True,
    daily_reports: bool = True,
    test_only: bool = False,
    filename_prefix: str = "",
) -> list[Path]:
    sheets = read_equipment_workbook(xlsx_path)
    created: list[Path] = []
    cards_dir = output_dir / "device-cards"
    checklists_dir = output_dir / "checklists"
    daily_dir = output_dir / "daily-reports"

    if test_only:
        created.append(write_open_me_first(output_dir / f"{filename_prefix}Open_Me_First.xlsx"))

    for sheet_name, raw_df in sheets.items():
        mapped = map_dataframe(raw_df)
        devices = iter_devices(mapped)
        if test_only:
            devices = devices[:1]

        for device in devices:
            slug = device_slug(device)
            if cards:
                card_path = cards_dir / f"{filename_prefix}Card_{slug}.xlsx"
                created.extend(write_device_card(card_path, device, template_dir))
            if checklists:
                checklist_path = checklists_dir / f"{filename_prefix}Checklist_{slug}.xlsx"
                created.extend(write_device_checklist(checklist_path, device, template_dir))
            if daily_reports:
                daily_path = daily_dir / f"{filename_prefix}DailyReport_{slug}.xlsx"
                created.extend(write_device_daily_report(daily_path, device, template_dir))

        if test_only and devices:
            bundle_path = output_dir / f"{filename_prefix}Sample_Device_Package.xlsx"
            write_device_sample_bundle(bundle_path, devices[0], template_dir)
            created.append(bundle_path)

            download_dir = output_dir / "download"
            download_files = write_download_test_package(download_dir, devices[0], template_dir)
            created.extend(download_files)
            break

    return created


def split_by_unit(df: pd.DataFrame, fallback_unit: str) -> dict[str, pd.DataFrame]:
    if df["unit"].str.strip().ne("").any():
        grouped: dict[str, pd.DataFrame] = {}
        for unit_name, group in df.groupby(df["unit"].str.strip(), dropna=False):
            key = str(unit_name).strip() or fallback_unit
            grouped[key] = group.reset_index(drop=True)
        return grouped

    return {fallback_unit: df.reset_index(drop=True)}


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", name.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned[:120] or "unit"


def style_cell(cell, *, bold: bool = False, fill: PatternFill | None = None, wrap: bool = False):
    cell.font = Font(name="Calibri", size=11, bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


def write_template(
    output_path: Path,
    unit_name: str,
    equipment_rows: pd.DataFrame,
    txt_context: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 20

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    title = ws.cell(row=row, column=1, value="ВЕДОМОСТЬ ЗАКРЕПЛЕНИЯ ОРГТЕХНИКИ ЗА ПОДРАЗДЕЛЕНИЕМ")
    style_cell(title, bold=True, fill=HEADER_FILL)
    title.alignment = Alignment(horizontal="center", vertical="center")

    row += 2
    meta = [
        ("Подразделение:", unit_name),
        ("Дата составления:", date.today().strftime("%d.%m.%Y")),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label)
        style_cell(ws.cell(row=row, column=1), bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2, value=value)
        style_cell(ws.cell(row=row, column=2))
        row += 1

    if txt_context.strip():
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        note = ws.cell(row=row, column=1, value="Дополнительная информация")
        style_cell(note, bold=True, fill=HEADER_FILL)
        row += 1

        paragraphs = [p.strip() for p in txt_context.split("\n\n") if p.strip()]
        for paragraph in paragraphs[:3]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(row=row, column=1, value=paragraph)
            style_cell(cell, wrap=True)
            row += 1

    row += 1
    headers = [
        "№",
        "Наименование",
        "Модель",
        "Серийный номер",
        "Инвентарный номер",
        "Местоположение",
        "Ответственный",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        style_cell(cell, bold=True, fill=TABLE_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    start_data_row = row

    for index, item in equipment_rows.iterrows():
        name = item.get("equipment_name", "") or item.get("model", "")
        values = [
            index + 1,
            name,
            item.get("model", ""),
            item.get("serial", ""),
            item.get("inventory", ""),
            item.get("location", ""),
            item.get("employee", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            style_cell(cell, wrap=True)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    if equipment_rows.empty:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value="Оборудование не найдено")
        style_cell(cell)
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="Передал: __________________ / __________________")
    style_cell(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(row=row, column=5, value="Принял: __________________ / __________________")
    style_cell(ws.cell(row=row, column=5))

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    footer = ws.cell(row=row, column=1, value="Подписи. М.П.")
    style_cell(footer)

    ws.freeze_panes = f"A{start_data_row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_templates(
    xlsx_path: Path,
    txt_dir: Path,
    output_dir: Path,
    sample_only: str | None = None,
    filename_prefix: str = "",
) -> list[Path]:
    txt_context = load_txt_context(txt_dir)
    sheets = read_equipment_workbook(xlsx_path)
    created: list[Path] = []

    for sheet_name, raw_df in sheets.items():
        mapped = map_dataframe(raw_df)
        units = split_by_unit(mapped, fallback_unit=sheet_name)

        for unit_name, unit_df in units.items():
            if sample_only and unit_name != sample_only and sheet_name != sample_only:
                continue

            filename = f"{filename_prefix}Ведомость_{sanitize_filename(unit_name)}.xlsx"
            output_path = output_dir / filename
            write_template(output_path, unit_name, unit_df, txt_context)
            created.append(output_path)

            if sample_only:
                return created

    return created


def find_source_workbook(explicit_path: Path | None, input_dir: Path) -> Path | None:
    if explicit_path and explicit_path.exists():
        return explicit_path

    if not input_dir.exists():
        return None

    candidates = sorted(input_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for candidate in candidates:
        if candidate.name.startswith("~$"):
            continue
        return candidate
    return None


def build_sample_data() -> tuple[pd.DataFrame, str]:
    equipment = pd.DataFrame(
        [
            {
                "equipment_name": "МФУ",
                "model": "HP LaserJet Pro M428fdw",
                "serial": "VNC3K12345",
                "inventory": "INV-2024-001",
                "location": "3 этаж, каб. 305",
                "employee": "Иванов И.И.",
                "unit": "Отдел продаж",
            },
            {
                "equipment_name": "Принтер",
                "model": "Canon imageRUNNER 2625i",
                "serial": "QWE987654",
                "inventory": "INV-2024-002",
                "location": "3 этаж, каб. 307",
                "employee": "Петрова А.С.",
                "unit": "Отдел продаж",
            },
            {
                "equipment_name": "Сканер",
                "model": "Epson WorkForce DS-770",
                "serial": "EPS112233",
                "inventory": "INV-2024-003",
                "location": "3 этаж, архив",
                "employee": "",
                "unit": "Отдел продаж",
            },
        ]
    )

    txt_context = (
        "Шаблон составлен для закрепления оргтехники за сотрудниками подразделения.\n"
        "Ответственные лица обязаны обеспечить сохранность оборудования и своевременную инвентаризацию."
    )
    return equipment, txt_context


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate personnel equipment templates")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("equipment-templates/input/Этажи покопийка AI.xlsx"),
        help="Source workbook with printers and office equipment",
    )
    parser.add_argument(
        "--txt-dir",
        type=Path,
        default=Path("equipment-templates/input/txt"),
        help="Directory with Qwen txt context files",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("equipment-templates/output"),
        help="Directory for generated templates",
    )
    parser.add_argument(
        "--sample-only",
        type=str,
        default=None,
        help="Generate only one unit (by unit or sheet name) for approval",
    )
    parser.add_argument(
        "--demo",
        action="store_true",
        help="Generate a demo sample without source files",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Print workbook sheets and detected columns, then exit",
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="Generate one unit vedemost template from the first department",
    )
    parser.add_argument(
        "--device-test",
        action="store_true",
        help="Generate one inventory card and one checklist for the first device",
    )
    parser.add_argument(
        "--device-all",
        action="store_true",
        help="Generate inventory cards, checklists, and daily reports for all devices",
    )
    parser.add_argument(
        "--templates-dir",
        type=Path,
        default=Path("equipment-templates/input"),
        help="Directory with Etagi txt templates (also searches input/templates/)",
    )
    args = parser.parse_args()

    input_dir = Path("equipment-templates/input")
    explicit = args.xlsx if args.xlsx.exists() else None
    source_xlsx = find_source_workbook(explicit, input_dir)

    if args.inspect:
        if not source_xlsx:
            raise SystemExit(f"No xlsx found in {input_dir}. Upload 'Этажи покопийка AI.xlsx' there.")
        sheets = read_equipment_workbook(source_xlsx)
        print(f"Source: {source_xlsx}")
        for sheet_name, raw_df in sheets.items():
            mapped = map_dataframe(raw_df)
            print(f"\nSheet: {sheet_name} ({len(mapped)} rows)")
            print("Columns:", ", ".join(raw_df.columns))
            print("Detected:", ", ".join(f"{k}={v or '-'}" for k, v in {
                "model": detect_column(list(raw_df.columns), "model"),
                "serial": detect_column(list(raw_df.columns), "serial"),
                "inventory": detect_column(list(raw_df.columns), "inventory"),
                "location": detect_column(list(raw_df.columns), "location"),
                "unit": detect_column(list(raw_df.columns), "unit"),
            }.items()))
        return

    if args.demo:
        output = args.output_dir / "SAMPLE_Ведомость_для_согласования.xlsx"
        equipment, txt_context = build_sample_data()
        write_template(output, "Отдел продаж (образец)", equipment, txt_context)
        print(f"Created demo sample: {output}")
        return

    if not source_xlsx:
        raise SystemExit(
            f"Source file not found.\n"
            f"Upload 'Этажи покопийка AI.xlsx' to: {input_dir.resolve()}\n"
            f"Then run: python equipment-templates/generate_templates.py --device-test"
        )

    if args.device_test or args.device_all:
        created = generate_device_documents(
            source_xlsx,
            args.output_dir,
            args.templates_dir,
            test_only=args.device_test,
            filename_prefix="TEST_" if args.device_test else "",
        )
        if not created:
            raise SystemExit("No device documents were generated.")
        for path in created:
            print(f"Created: {path}")
        if args.device_test:
            download_zip = args.output_dir / "download" / "TEST_Package.zip"
            print("\nDownload from GitHub:")
            print(f"  {download_zip}")
            print("Or open: equipment-templates/output/download/TEST_All_In_One.xlsx")
        return

    sample_only = args.sample_only
    if args.test and not sample_only:
        sheets = read_equipment_workbook(source_xlsx)
        first_sheet = next(iter(sheets))
        mapped = map_dataframe(sheets[first_sheet])
        mapped = mapped[mapped["unit"].str.strip().ne("")].reset_index(drop=True)
        units = split_by_unit(mapped, first_sheet)
        sample_only = next(
            (name for name in units if name.strip() and name != first_sheet),
            next(iter(units)),
        )
        args.output_dir.mkdir(parents=True, exist_ok=True)

    prefix = "TEST_" if args.test else ""
    created = generate_templates(
        source_xlsx, args.txt_dir, args.output_dir, sample_only, filename_prefix=prefix
    )
    if not created:
        raise SystemExit("No templates were generated. Check workbook sheets and columns.")

    for path in created:
        print(f"Created: {path}")


if __name__ == "__main__":
    main()
