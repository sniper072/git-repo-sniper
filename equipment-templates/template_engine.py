"""Load Etagi txt templates and render them into Excel without splitting text."""

from __future__ import annotations

import re
import zipfile
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

INPUT_DIR = Path("equipment-templates/input")
TEMPLATE_DIR = INPUT_DIR / "templates"

TEMPLATE_FILES = {
    "card": ["Etagi_equipment_Card.txt", "Etagi_equipment_card.txt"],
    "checklist": ["Etagi_checklist_template.txt"],
    "daily_report": ["Etagi_daily_report.txt", "Etagi_ daily_report.txt"],
}

OUTPUT_COLUMNS = ["№", "Поле", "Значение", "Ед.", "Отметка"]
LEGACY_COLUMNS = ["Поле", "Значение", "Ед.", "Отметка"]

FILL_HEADER = PatternFill("solid", fgColor="2F5496")
FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")
FILL_STEP = PatternFill("solid", fgColor="E2EFDA")
FILL_LABEL = PatternFill("solid", fgColor="F2F2F2")
FILL_AUTO = PatternFill("solid", fgColor="DDEBF7")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FILL_MARK = PatternFill("solid", fgColor="FFE699")
FILL_TEXT = PatternFill("solid", fgColor="FFFFFF")
FILL_NOTE = PatternFill("solid", fgColor="EDEDED")
FILL_DEVICE_STRIP = PatternFill("solid", fgColor="E2EFDA")
FILL_ALERT = PatternFill("solid", fgColor="FCE4D6")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="1F3864")
FONT_STEP = Font(name="Calibri", size=10, bold=True, color="375623")
FONT_LABEL = Font(name="Calibri", size=10, bold=True)
FONT_BODY = Font(name="Calibri", size=10)
FONT_HINT = Font(name="Calibri", size=9, italic=True, color="666666")
FONT_NUM = Font(name="Calibri", size=10, bold=True, color="1F3864")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
INPUT_BORDER = Border(
    left=Side(style="medium", color="BF8F00"),
    right=Side(style="medium", color="BF8F00"),
    top=Side(style="medium", color="BF8F00"),
    bottom=Side(style="medium", color="BF8F00"),
)
MARK_BORDER = Border(
    left=Side(style="medium", color="C65911"),
    right=Side(style="medium", color="C65911"),
    top=Side(style="medium", color="C65911"),
    bottom=Side(style="medium", color="C65911"),
)

DEFAULT_LAYOUT = "field"

# Labels from Etagi templates -> device field keys
UNDERSCORE_LABELS: list[tuple[str, str]] = [
    ("Модель:", "model"),
    ("Серийный номер:", "serial"),
    ("Инвентарный номер:", "inventory"),
    ("Локация (офис/кабинет):", "location"),
    ("Ответственный у Заказчика:", "employee"),
    ("Общий пробег (Total):", "report_counter"),
    ("Пробег с последнего ТО:", "print_count"),
]


def template_search_dirs(template_dir: Path | None = None) -> list[Path]:
    dirs: list[Path] = []
    if template_dir is not None:
        dirs.append(template_dir)
    dirs.extend([INPUT_DIR, TEMPLATE_DIR])
    unique: list[Path] = []
    for path in dirs:
        if path not in unique:
            unique.append(path)
    return unique


def find_template(kind: str, template_dir: Path | None = None) -> Path:
    for directory in template_search_dirs(template_dir):
        for name in TEMPLATE_FILES[kind]:
            path = directory / name
            if path.exists():
                return path
        if kind == "card":
            matches = sorted(directory.glob("*Card*.txt"))
        elif kind == "checklist":
            matches = sorted(directory.glob("*checklist*.txt"))
        else:
            matches = sorted(directory.glob("*daily*report*.txt"))
        if matches:
            return matches[0]

    expected = ", ".join(TEMPLATE_FILES[kind])
    raise FileNotFoundError(
        f"Template not found for '{kind}'.\n"
        f"Upload one of: {expected}\n"
        f"To: equipment-templates/input/"
    )


def load_template(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace").replace("\r\n", "\n")


def safe_text(value: object) -> str:
    return str(value or "").strip()


def build_context(device: pd.Series) -> dict[str, str]:
    location = safe_text(device.get("location", ""))
    office = safe_text(device.get("office", ""))
    unit = safe_text(device.get("unit", ""))
    if location and office and office not in location:
        location_display = f"{location} / {office}"
    elif location:
        location_display = location
    elif office:
        location_display = office
    else:
        location_display = unit

    values = {
        "date": date.today().strftime("%d.%m.%Y"),
        "row_number": safe_text(device.get("row_number", "")),
        "inventory": safe_text(device.get("inventory", "")),
        "model": safe_text(device.get("model", "")),
        "serial": safe_text(device.get("serial", "")),
        "location": location_display,
        "office": office,
        "unit": unit,
        "employee": safe_text(device.get("employee", "")),
        "prev_counter": safe_text(device.get("prev_counter", "")),
        "report_counter": safe_text(device.get("report_counter", "")),
        "print_count": safe_text(device.get("print_count", "")),
    }

    context: dict[str, str] = {}
    for key, value in values.items():
        context[key] = value
        context[key.upper()] = value

    russian = {
        "дата": values["date"],
        "модель": values["model"],
        "серийный номер": values["serial"],
        "инвентарный номер": values["inventory"],
        "местоположение": values["location"],
    }
    context.update(russian)
    return context


def replace_underscores_after_label(line: str, label: str, value: str) -> str:
    if label not in line or not value:
        return line

    label_index = line.find(label)
    before = line[: label_index + len(label)]
    after = line[label_index + len(label) :]
    match = re.match(r"(\s*)(_+)(.*)$", after)
    if not match:
        return line

    spaces, underscores, suffix = match.groups()
    replacement = value[: len(underscores)].ljust(len(underscores))
    return f"{before}{spaces}{replacement}{suffix}"


def fill_special_patterns(text: str, context: dict[str, str]) -> str:
    text = re.sub(
        r"КАРТОЧКА АППАРАТА №\s*_+",
        f"КАРТОЧКА АППАРАТА № {context['inventory'] or '_______'}",
        text,
        count=1,
    )
    text = re.sub(
        r"(?m)(Дата:\s*)_+",
        rf"\g<1>{context['date']}",
        text,
    )
    text = re.sub(
        r"(?m)^(   )_+\s*$",
        rf"\g<1>{context['location']}",
        text,
        count=1,
    )
    return text


def fill_underscore_fields(text: str, context: dict[str, str]) -> str:
    lines = []
    for line in text.split("\n"):
        updated = line
        for label, field_key in UNDERSCORE_LABELS:
            updated = replace_underscores_after_label(updated, label, context.get(field_key, ""))
        lines.append(updated)
    return "\n".join(lines)


def render_template(template_text: str, device: pd.Series) -> str:
    context = build_context(device)
    rendered = template_text

    for key, value in sorted(context.items(), key=lambda item: len(item[0]), reverse=True):
        rendered = rendered.replace(f"{{{{{key}}}}}", value)
        rendered = rendered.replace(f"{{{key}}}", value)

    rendered = fill_special_patterns(rendered, context)
    rendered = fill_underscore_fields(rendered, context)
    return rendered


def validate_xlsx(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    workbook.close()


def is_decorative_line(stripped: str) -> bool:
    if not stripped:
        return True
    if re.fullmatch(r"[═─_\-]+", stripped):
        return True
    if stripped.startswith("┌") or stripped.startswith("└"):
        return True
    return False


def strip_decorative_chars(text: str) -> str:
    if not text:
        return ""
    cleaned = text.replace("☐", "").replace("□", "")
    if cleaned.startswith("│") and cleaned.endswith("│"):
        cleaned = cleaned[1:-1]
    cleaned = cleaned.strip(" │═─-_")
    return cleaned.strip()


def parse_rendered_line(line: str) -> tuple[str, str, str, str, str]:
    """Split one template line into: type, field, value, unit, mark."""
    raw = line.rstrip()
    stripped = raw.strip()

    if not stripped or is_decorative_line(stripped):
        return ("", "", "", "", "")

    if stripped.startswith("│") and stripped.endswith("│"):
        title = strip_decorative_chars(stripped)
        if title:
            return ("Раздел", title, "", "", "")
        return ("", "", "", "", "")

    if re.match(r"^\d+\.\s+", stripped):
        return ("Раздел", strip_decorative_chars(stripped), "", "", "")

    if re.match(r"^ШАГ\s+\d+", stripped, flags=re.IGNORECASE):
        return ("Шаг", strip_decorative_chars(stripped), "", "", "")

    checkbox_match = re.match(r"^(\[\s*\]|\[.\]|□)\s*(.+)$", stripped)
    if checkbox_match:
        return ("Проверка", strip_decorative_chars(checkbox_match.group(2)), "", "", "")

    rating_match = re.match(r"^(.+?):\s*((?:\[\s*\d\s*\]\s*)+)$", stripped)
    if rating_match:
        return ("Оценка", strip_decorative_chars(rating_match.group(1)), "", "", rating_match.group(2).strip())

    if "🟢" in stripped or "🟡" in stripped or "🔴" in stripped:
        label_match = re.match(r"^([🟢🟡🔴]\s*.+?:)\s*(.*)$", stripped)
        if label_match:
            value = label_match.group(2).strip()
            value = "" if re.fullmatch(r"_+\s*шт\.?", value) else value.replace("_", "").strip()
            return ("Поле", strip_decorative_chars(label_match.group(1)), value, "", "")

    colon_match = re.match(r"^(.+?):\s*(.*)$", stripped)
    if colon_match:
        label = strip_decorative_chars(colon_match.group(1))
        value_raw = colon_match.group(2).strip()
        if label.startswith("КАРТОЧКА АППАРАТА №"):
            inventory = label.replace("КАРТОЧКА АППАРАТА №", "").strip()
            if inventory:
                return ("Поле", "КАРТОЧКА АППАРАТА №", inventory, "", "")
        value, unit = split_counter_value(value_raw)
        return ("Поле", label, value, unit, "")

    if stripped.startswith("ПОДПИСИ") or stripped.startswith("Антон:") or stripped.startswith("Проверил:"):
        return ("Подпись", strip_decorative_chars(stripped), "", "", "")

    if stripped.startswith("Комментарий") or stripped.startswith("Краткое описание"):
        return ("Поле", strip_decorative_chars(stripped), "", "", "")

    if set(stripped) <= {"_", "─", "═", "-"}:
        return ("", "", "", "", "")

    cleaned = strip_decorative_chars(stripped)
    if not cleaned or is_decorative_line(cleaned):
        return ("", "", "", "", "")

    return ("Текст", cleaned, "", "", "")


def split_counter_value(value_raw: str) -> tuple[str, str]:
    text = clean_value(value_raw)
    if not text:
        unit_match = re.search(r"(стр\.|шт\.)", value_raw)
        return "", unit_match.group(1) if unit_match else ""

    match = re.match(r"^([\d\s]+)\s*(стр\.|шт\.)?$", text)
    if match:
        return match.group(1).replace(" ", ""), match.group(2) or ""

    match = re.match(r"^(\d+)\s*(стр\.|шт\.)$", text)
    if match:
        return match.group(1), match.group(2)

    # value with trailing unit attached
    parts = text.rsplit(" ", 1)
    if len(parts) == 2 and parts[1] in {"стр.", "шт."}:
        return parts[0], parts[1]

    return text, ""


def clean_value(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    if re.fullmatch(r"_+", text):
        return ""
    text = re.sub(r"^_+\s*", "", text)
    text = re.sub(r"\s*_+$", "", text)
    if text in {"стр.", "шт."}:
        return ""
    return text.strip()


def expand_line_rows(line: str) -> list[tuple[str, str, str, str, str]]:
    stripped = line.strip()
    if "Дата:" in stripped and "Инженер:" in stripped:
        rows: list[tuple[str, str, str, str, str]] = []
        for label in ("Дата", "Инженер"):
            match = re.search(rf"{label}:\s*([^:]+?)(?=\s+(?:Дата|Инженер):|$)", stripped)
            if match:
                rows.append(("Поле", label, clean_value(match.group(1)), "", ""))
        if rows:
            return rows

    if stripped.startswith("КАРТОЧКА АППАРАТА") and "Дата:" in stripped:
        rows = []
        inv_match = re.search(r"КАРТОЧКА АППАРАТА №\s*(\S+)", stripped)
        date_match = re.search(r"Дата:\s*(\S+)", stripped)
        if inv_match:
            rows.append(("Поле", "КАРТОЧКА АППАРАТА №", inv_match.group(1), "", ""))
        if date_match:
            rows.append(("Поле", "Дата", date_match.group(1), "", ""))
        if rows:
            return rows

    return [parse_rendered_line(line)]


def rendered_text_to_dataframe(rendered_text: str) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    pending_location = False

    for line in rendered_text.split("\n"):
        stripped = line.strip()

        if pending_location and line.startswith("   ") and stripped and not stripped.startswith("["):
            rows.append(
                {
                    "Поле": "Локация дня",
                    "Значение": stripped,
                    "Ед.": "",
                    "Отметка": "",
                    "row_kind": "Поле",
                }
            )
            pending_location = False
            continue

        for row_type, field, value, unit, mark in expand_line_rows(line):
            if not any([row_type, field, value, unit, mark]):
                continue
            if row_type == "Раздел" and "ЛОКАЦИЯ ДНЯ" in field.upper():
                pending_location = True
            else:
                pending_location = False
            rows.append(
                {
                    "Поле": field,
                    "Значение": value,
                    "Ед.": unit,
                    "Отметка": mark,
                    "row_kind": row_type,
                }
            )

    if not rows:
        rows.append({"Поле": "", "Значение": "", "Ед.": "", "Отметка": "", "row_kind": ""})

    return pd.DataFrame(rows, columns=[*OUTPUT_COLUMNS, "row_kind"]).fillna("")


def output_frame(frame: pd.DataFrame) -> pd.DataFrame:
    numbered = number_check_rows(frame)
    return numbered[OUTPUT_COLUMNS].astype(str).replace("nan", "")


def number_check_rows(frame: pd.DataFrame) -> pd.DataFrame:
    """Add № column with sequential numbers for checklist rows."""
    result = frame.copy()
    result["№"] = ""
    check_no = 0
    for index, row in result.iterrows():
        if str(row.get("row_kind", "")) == "Проверка":
            check_no += 1
            result.at[index, "№"] = str(check_no)
    return result


def extract_device_summary(frame: pd.DataFrame) -> dict[str, str]:
    labels = {
        "Модель": "model",
        "Серийный номер": "serial",
        "Инвентарный номер": "inventory",
        "Локация (офис/кабинет)": "location",
        "Локация дня": "location",
    }
    summary = {key: "" for key in labels.values()}
    for row in frame.itertuples(index=False):
        field = str(getattr(row, "Поле", "") or "")
        value = str(getattr(row, "Значение", "") or "")
        if field in labels and value and value not in {"", "nan"}:
            summary[labels[field]] = value
    return summary


def is_alert_section(field: str) -> bool:
    upper = field.upper()
    return "КРАСНЫЕ ФЛАГИ" in upper or "ЭСКАЛАЦ" in upper


def is_title_row(field: str, kind: str, row_index: int) -> bool:
    upper = field.upper()
    if row_index == 0 and kind in {"Раздел", "Текст"}:
        return any(token in upper for token in ("КАРТОЧКА", "ЧЕК-ЛИСТ", "ЕЖЕДНЕВНЫЙ", "ОТЧЕТ"))
    return False


def set_cell_style(
    cell,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = THIN_BORDER,
) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def iter_form_rows(frame: pd.DataFrame):
    """Yield row dicts; avoid itertuples (breaks on «Ед.» and «№» column names)."""
    for _, row in frame.iterrows():
        yield {
            "field": str(row.get("Поле", "") or ""),
            "value": str(row.get("Значение", "") or ""),
            "unit": str(row.get("Ед.", "") or ""),
            "mark": str(row.get("Отметка", "") or ""),
            "kind": str(row.get("row_kind", "") or ""),
            "num": str(row.get("№", "") or ""),
        }


def apply_form_layout(worksheet, frame: pd.DataFrame, layout_mode: str = DEFAULT_LAYOUT) -> None:
    """Apply entry-friendly Excel layout. Modes: 'field' (numbered, device strip) or 'compact'."""
    data = number_check_rows(frame.fillna(""))
    use_numbering = layout_mode != "compact"
    columns = OUTPUT_COLUMNS if use_numbering else LEGACY_COLUMNS
    last_col = len(columns)
    summary = extract_device_summary(frame)
    check_counter = 0

    worksheet.sheet_view.showGridLines = False
    if use_numbering:
        worksheet.column_dimensions["A"].width = 5
        worksheet.column_dimensions["B"].width = 42
        worksheet.column_dimensions["C"].width = 26
        worksheet.column_dimensions["D"].width = 7
        worksheet.column_dimensions["E"].width = 14
    else:
        worksheet.column_dimensions["A"].width = 46
        worksheet.column_dimensions["B"].width = 24
        worksheet.column_dimensions["C"].width = 7
        worksheet.column_dimensions["D"].width = 16

    headers = (
        ["№", "Поле", "Значение", "Ед.", "Отметка ✓/✗"]
        if use_numbering
        else ["Поле", "Значение", "Ед.", "Отметка ✓/✗"]
    )
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        set_cell_style(
            cell,
            font=FONT_HEADER,
            fill=FILL_HEADER,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )
    worksheet.row_dimensions[1].height = 24

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    legend_text = (
        "№ = пункт проверки  |  Жёлтые = ввод  |  Голубые = из базы  |  Отметка: ✓ / ✗ / 1-5"
        if use_numbering
        else "Жёлтые ячейки — ввод вручную  |  Голубые — из базы  |  Отметка: ✓ / ✗ / 1-5"
    )
    legend = worksheet.cell(row=2, column=1, value=legend_text)
    set_cell_style(
        legend,
        font=FONT_HINT,
        fill=FILL_NOTE,
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=None,
    )
    worksheet.row_dimensions[2].height = 18

    excel_row = 3
    if use_numbering and any(summary.values()):
        strip_parts = []
        if summary["model"]:
            strip_parts.append(f"Модель: {summary['model']}")
        if summary["serial"]:
            strip_parts.append(f"Серийный: {summary['serial']}")
        if summary["inventory"]:
            strip_parts.append(f"Инв. {summary['inventory']}")
        if summary["location"]:
            strip_parts.append(f"Локация: {summary['location']}")
        worksheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=last_col)
        strip_cell = worksheet.cell(row=excel_row, column=1, value="  |  ".join(strip_parts))
        set_cell_style(
            strip_cell,
            font=FONT_LABEL,
            fill=FILL_DEVICE_STRIP,
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=THIN_BORDER,
        )
        worksheet.row_dimensions[excel_row].height = 22
        excel_row += 1

    mark_rows: list[str] = []
    rating_rows: list[str] = []
    prev_kind = ""

    def col(name: str) -> int:
        mapping = {"num": 1, "field": 2, "value": 3, "unit": 4, "mark": 5}
        if not use_numbering:
            mapping = {"field": 1, "value": 2, "unit": 3, "mark": 4}
        return mapping[name]

    for offset, item in enumerate(iter_form_rows(data)):
        field = item["field"]
        value = item["value"]
        unit = item["unit"]
        mark = item["mark"]
        kind = item["kind"]
        num = item["num"]

        if use_numbering and kind == "Раздел" and prev_kind and prev_kind not in {"", "Раздел", "Шаг"}:
            worksheet.row_dimensions[excel_row].height = 6
            excel_row += 1

        if kind == "Поле" and field == "КАРТОЧКА АППАРАТА №" and value:
            worksheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=last_col)
            title_cell = worksheet.cell(row=excel_row, column=1, value=f"КАРТОЧКА АППАРАТА № {value}")
            set_cell_style(
                title_cell,
                font=FONT_TITLE,
                fill=FILL_TITLE,
                alignment=Alignment(horizontal="center", vertical="center"),
            )
            worksheet.row_dimensions[excel_row].height = 28
            excel_row += 1
            prev_kind = kind
            continue

        if is_title_row(field, kind, offset):
            worksheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=last_col)
            title_cell = worksheet.cell(row=excel_row, column=1, value=field)
            set_cell_style(
                title_cell,
                font=FONT_TITLE,
                fill=FILL_TITLE,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            )
            worksheet.row_dimensions[excel_row].height = 28
            excel_row += 1
            prev_kind = kind
            continue

        section_fill = FILL_ALERT if is_alert_section(field) else FILL_SECTION
        if kind == "Раздел" or (kind == "Текст" and field.isupper() and len(field) > 20):
            worksheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=last_col)
            section_cell = worksheet.cell(row=excel_row, column=1, value=field)
            set_cell_style(
                section_cell,
                font=FONT_SECTION,
                fill=section_fill,
                alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            )
            worksheet.row_dimensions[excel_row].height = 22
            excel_row += 1
            prev_kind = kind
            continue

        if kind == "Шаг":
            worksheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=last_col)
            step_cell = worksheet.cell(row=excel_row, column=1, value=field)
            set_cell_style(
                step_cell,
                font=FONT_STEP,
                fill=FILL_STEP,
                alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            )
            worksheet.row_dimensions[excel_row].height = 20
            excel_row += 1
            prev_kind = kind
            continue

        if use_numbering:
            num_cell = worksheet.cell(row=excel_row, column=col("num"), value=num if num else "")
            set_cell_style(
                num_cell,
                font=FONT_NUM,
                fill=FILL_NOTE,
                alignment=Alignment(horizontal="center", vertical="center"),
            )

        label_cell = worksheet.cell(row=excel_row, column=col("field"), value=field)
        value_cell = worksheet.cell(
            row=excel_row, column=col("value"), value="" if value in {"", "nan"} else value
        )
        unit_cell = worksheet.cell(row=excel_row, column=col("unit"), value="" if unit in {"", "nan"} else unit)
        mark_cell = worksheet.cell(row=excel_row, column=col("mark"), value="" if mark in {"", "nan"} else mark)

        set_cell_style(
            label_cell,
            font=FONT_LABEL if kind == "Поле" else FONT_BODY,
            fill=FILL_LABEL if kind == "Поле" else FILL_TEXT,
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )
        set_cell_style(
            unit_cell,
            font=FONT_BODY,
            fill=FILL_NOTE,
            alignment=Alignment(horizontal="center", vertical="center"),
        )

        mark_col_letter = "E" if use_numbering else "D"
        row_ref = f"{mark_col_letter}{excel_row}"

        if kind == "Проверка":
            check_counter += 1
            if use_numbering:
                num_cell.value = str(check_counter)
            set_cell_style(
                label_cell,
                font=FONT_BODY,
                fill=FILL_TEXT,
                alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
            )
            set_cell_style(value_cell, font=FONT_BODY, fill=FILL_TEXT, alignment=Alignment(horizontal="left", vertical="center"))
            set_cell_style(
                mark_cell,
                font=FONT_BODY,
                fill=FILL_MARK,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=MARK_BORDER,
            )
            mark_rows.append(row_ref)
            worksheet.row_dimensions[excel_row].height = max(22, min(50, 16 + len(field) // 45 * 6))
            excel_row += 1
            prev_kind = kind
            continue

        if kind == "Оценка":
            if mark and "[" in mark:
                mark_cell.value = ""
            set_cell_style(value_cell, font=FONT_BODY, fill=FILL_TEXT, alignment=Alignment(horizontal="left", vertical="center"))
            set_cell_style(
                mark_cell,
                font=FONT_BODY,
                fill=FILL_MARK,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=MARK_BORDER,
            )
            rating_rows.append(row_ref)
            worksheet.row_dimensions[excel_row].height = 22
            excel_row += 1
            prev_kind = kind
            continue

        if kind == "Подпись":
            worksheet.merge_cells(
                start_row=excel_row, start_column=col("field"), end_row=excel_row, end_column=col("value")
            )
            sign_col = col("unit")
            worksheet.merge_cells(
                start_row=excel_row, start_column=sign_col, end_row=excel_row, end_column=last_col
            )
            set_cell_style(label_cell, font=FONT_LABEL, fill=FILL_LABEL, alignment=Alignment(horizontal="left", vertical="center"))
            sign_cell = worksheet.cell(row=excel_row, column=sign_col, value="подпись / дата")
            set_cell_style(
                sign_cell,
                font=FONT_HINT,
                fill=FILL_INPUT,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=INPUT_BORDER,
            )
            worksheet.row_dimensions[excel_row].height = 26
            excel_row += 1
            prev_kind = kind
            continue

        value_empty = value in {"", "nan"}
        is_counter = unit in {"стр.", "шт."}
        is_comment = field.startswith("Комментарий") or field.startswith("Краткое описание")

        if is_comment or (kind == "Текст" and not value):
            worksheet.merge_cells(
                start_row=excel_row, start_column=col("value"), end_row=excel_row, end_column=last_col
            )
            set_cell_style(
                value_cell,
                font=FONT_BODY,
                fill=FILL_INPUT,
                alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                border=INPUT_BORDER,
            )
            worksheet.row_dimensions[excel_row].height = 32
            excel_row += 1
            prev_kind = kind
            continue

        value_fill = FILL_INPUT if value_empty or is_counter else FILL_AUTO
        value_border = INPUT_BORDER if value_empty or is_counter else THIN_BORDER
        set_cell_style(
            value_cell,
            font=FONT_BODY,
            fill=value_fill,
            alignment=Alignment(
                horizontal="right" if is_counter and not value_empty else "left",
                vertical="center",
                wrap_text=True,
            ),
            border=value_border,
        )
        set_cell_style(
            mark_cell,
            font=FONT_BODY,
            fill=FILL_TEXT,
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        worksheet.row_dimensions[excel_row].height = 22
        excel_row += 1
        prev_kind = kind

    if mark_rows:
        mark_validation = DataValidation(
            type="list",
            formula1='"✓,✗,—"',
            allow_blank=True,
            showDropDown=True,
        )
        worksheet.add_data_validation(mark_validation)
        for cell_ref in mark_rows:
            mark_validation.add(cell_ref)

    if rating_rows:
        rating_validation = DataValidation(
            type="list",
            formula1='"1,2,3,4,5"',
            allow_blank=True,
            showDropDown=True,
        )
        worksheet.add_data_validation(rating_validation)
        for cell_ref in rating_rows:
            rating_validation.add(cell_ref)

    freeze_row = 4 if use_numbering and any(summary.values()) else 3
    worksheet.freeze_panes = f"A{freeze_row}"
    worksheet.print_title_rows = "1:2"
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True


def score_layout_workbook(path: Path) -> dict[str, int]:
    """Score a generated form for manual-entry usability (higher = better)."""
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook.active
    scores = {
        "input_cells": 0,
        "marked_checks": 0,
        "numbered_checks": 0,
        "device_strip": 0,
        "validations": 0,
        "section_gaps": 0,
        "total": 0,
    }

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        for cell in row:
            fill_rgb = getattr(getattr(cell.fill, "fgColor", None), "rgb", "") or ""
            value = str(cell.value or "")
            if fill_rgb in {"00FFF2CC", "00FFE699"}:
                scores["input_cells"] += 1
            if fill_rgb == "00FFE699" and cell.column in {4, 5}:
                scores["marked_checks"] += 1
            if cell.column == 1 and value.isdigit():
                scores["numbered_checks"] += 1
            if "Модель:" in value and "Серийный:" in value:
                scores["device_strip"] = 25
            if worksheet.row_dimensions[cell.row].height and worksheet.row_dimensions[cell.row].height <= 8:
                scores["section_gaps"] += 1

    scores["validations"] = len(worksheet.data_validations.dataValidation) * 10
    scores["total"] = (
        scores["input_cells"] * 2
        + scores["marked_checks"] * 3
        + scores["numbered_checks"] * 4
        + scores["device_strip"]
        + scores["validations"]
        + scores["section_gaps"] * 2
    )
    workbook.close()
    return scores


def compare_layouts_on_devices(
    devices: list[pd.Series],
    template_dir: Path,
    output_dir: Path,
) -> tuple[str, dict]:
    """Generate compact vs field layouts for each device and return winning mode."""
    output_dir.mkdir(parents=True, exist_ok=True)
    totals = {"compact": 0, "field": 0}
    details: list[str] = []

    for index, device in enumerate(devices, start=1):
        card_text = render_template(load_template(find_template("card", template_dir)), device)
        frame = rendered_text_to_dataframe(card_text)
        for mode in ("compact", "field"):
            path = output_dir / f"COMPARE_device{index}_{mode}_Card.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Card"
            apply_form_layout(worksheet, frame, layout_mode=mode)
            workbook.save(path)
            scores = score_layout_workbook(path)
            totals[mode] += scores["total"]
            details.append(
                f"Device {index} ({device.get('model', '?')}) / {mode}: "
                f"score={scores['total']} (inputs={scores['input_cells']}, "
                f"numbered={scores['numbered_checks']}, strip={scores['device_strip']})"
            )

    winner = "field" if totals["field"] >= totals["compact"] else "compact"
    report = "\n".join(details) + f"\n\nTOTAL compact={totals['compact']} field={totals['field']} → winner: {winner}"
    (output_dir / "LAYOUT_COMPARE_REPORT.txt").write_text(report, encoding="utf-8")
    return winner, totals


def write_dataframe_xlsx(path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    """Write dataframe with openpyxl and entry-friendly form layout."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    apply_form_layout(worksheet, frame)
    workbook.save(path)
    validate_xlsx(path)


def write_template_document(output_path: Path, sheet_name: str, rendered_text: str) -> Path:
    """Write template as columns: one value per cell for manual filling."""
    frame = rendered_text_to_dataframe(rendered_text)
    write_dataframe_xlsx(output_path, sheet_name, frame)
    return output_path


def write_template_bundle(output_path: Path, documents: dict[str, str]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rendered_text in documents.items():
        frame = rendered_text_to_dataframe(rendered_text)
        safe_name = sheet_name[:31]
        worksheet = workbook.create_sheet(title=safe_name)
        apply_form_layout(worksheet, frame)

    workbook.save(output_path)
    validate_xlsx(output_path)
    return output_path


def write_csv(path: Path, frame: pd.DataFrame) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    output_frame(frame).to_csv(path, index=False, encoding="utf-8-sig")
    return path


def create_zip(zip_path: Path, files: list[Path]) -> Path:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_path in files:
            if file_path.exists() and file_path.is_file():
                archive.write(file_path, arcname=file_path.name)
    return zip_path


def write_download_test_package(
    download_dir: Path,
    device: pd.Series,
    template_dir: Path,
) -> list[Path]:
    """Create simple ASCII-named files and one zip for easy GitHub download."""
    download_dir.mkdir(parents=True, exist_ok=True)

    card_text = render_template(load_template(find_template("card", template_dir)), device)
    checklist_text = render_template(load_template(find_template("checklist", template_dir)), device)
    daily_text = render_template(load_template(find_template("daily_report", template_dir)), device)

    documents = {
        "Card": card_text,
        "Checklist": checklist_text,
        "DailyReport": daily_text,
    }

    created: list[Path] = []
    simple_names = {
        "Card": "TEST_Card",
        "Checklist": "TEST_Checklist",
        "DailyReport": "TEST_Daily_Report",
    }

    for sheet_key, rendered in documents.items():
        base = simple_names[sheet_key]
        xlsx_path = download_dir / f"{base}.xlsx"
        csv_path = download_dir / f"{base}.csv"
        write_template_document(xlsx_path, sheet_key, rendered)
        write_csv(csv_path, rendered_text_to_dataframe(rendered))
        created.extend([xlsx_path, csv_path])

    all_in_one = download_dir / "TEST_All_In_One.xlsx"
    write_template_bundle(all_in_one, documents)
    created.append(all_in_one)

    readme = download_dir / "README.txt"
    readme.write_text(
        "DOWNLOAD TEST FILES\n"
        "===================\n\n"
        "Easiest: download TEST_Package.zip (all files in one archive).\n\n"
        "Single files in this folder:\n"
        "  TEST_All_In_One.xlsx  - card + checklist + daily report (3 sheets)\n"
        "  TEST_Card.xlsx\n"
        "  TEST_Checklist.xlsx\n"
        "  TEST_Daily_Report.xlsx\n\n"
        "CSV copies are included if Excel does not open.\n\n"
        "See LAYOUT_EXAMPLE.txt for color legend and entry workflow.\n",
        encoding="utf-8",
    )
    created.append(readme)

    layout_guide = download_dir / "LAYOUT_EXAMPLE.txt"
    if not layout_guide.exists():
        layout_guide.write_text(
            "See equipment-templates/output/download/LAYOUT_EXAMPLE.txt in repo.\n",
            encoding="utf-8",
        )
    created.append(layout_guide)

    zip_path = download_dir / "TEST_Package.zip"
    create_zip(zip_path, created)
    created.append(zip_path)
    return created
